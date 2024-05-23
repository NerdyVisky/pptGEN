import os
from langsmith import Client
from openpyxl import Workbook
from datetime import datetime
from langsmith import Client



client = Client()
PROJECT_NAME = os.environ['LANGCHAIN_PROJECT']
root_runs = client.list_runs(
    project_name=PROJECT_NAME,
    is_root=True
)

wb = Workbook()
ws = wb.active
ws.title = "API Log"

headers = ["Run ID", "Generation Name", "Day", "Start Time", "Execution time", "Total Cost", "Error"]
for col_num, header in enumerate(headers, start=1):
    ws.cell(row=1, column=col_num).value = header

row_num = 2 

for run in list(root_runs):
    start_time = run.start_time
    end_time = run.end_time
    total_cost = run.total_cost
    error = run.error
    # start_time = datetime.strptime(start_time_str, "%Y-%m-%dT%H:%M:%S.%f")
    # end_time = datetime.strptime(end_time_str, "%Y-%m-%dT%H:%M:%S.%f")
    
    # Calculate time taken in seconds
    time_taken = (end_time - start_time).total_seconds()
    day_of_exec = start_time.strftime("%d_%m")
    time_of_exec = start_time.strftime("%H_%M_%S")
    
    
    # Write data to the worksheet
    ws.cell(row=row_num, column=1).value = str(run.id)
    ws.cell(row=row_num, column=2).value = run.name
    ws.cell(row=row_num, column=3).value = day_of_exec
    ws.cell(row=row_num, column=4).value = time_of_exec
    ws.cell(row=row_num, column=5).value = time_taken
    ws.cell(row=row_num, column=6).value = total_cost
    ws.cell(row=row_num, column=7).value = error
    
    row_num += 1

# Save the workbook
now = datetime.now()
formatted_now = now.strftime("%H_%M_%d_%m")
if not os.path.exists('code\logs\\api_traces'):
    os.mkdir('code\logs\\api_traces')

wb.save(os.path.join('code\logs\\api_traces', f"{formatted_now}.xlsx"))
print(f"🟢 (5/5) API traces saved as {formatted_now}.xlsx in logs\\api_traces")

client.delete_project(project_name="pptGEN")